#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Auto Stock Report generator  (CORRECTED + EXECUTIVE DASHBOARD)
==============================================================
Reads 5 SAP exports (MB52 x2, R138 x2, Product Group) and produces a
multi-sheet Excel report:

  * "Executive Dashboard" - KPI cards + charts for management
  * "PV DATA (Dashboard)" - Plant / Product Group / Ageing / Client tables
  * "DATA"                - the cleaned row-level data
  * "Ageing > 365 D"      - dead-stock rows
  * copies of the 5 input sheets

Key logic (matches the manual report)
--------------------------------------
* Ageing      : measured as (today - Last GR), bucketed 0-30 / 31-90 / 91-180 /
                181-365 / >365.
* GR / Shipper / Profit center : tiered lookup against R138 so a row is matched
                even when MB52 'Unrestricted' differs from R138 'Quantity'
                (full key Material+Quantity+Batch -> Material+Batch -> Material).
* Product Group: Product Group file (by Material) first, then the RAW R138
                'Level 4 Product Group' (by Material) as the only fallback.
                Values are kept exactly as-is (e.g. 'SERVICE' from the Product
                Group file and 'SERVICES' from R138 remain distinct), which is
                how the manual report classifies them.

Everything is derived from the 5 daily input files only - no external/static
supplement, so the report is fully reproducible each day.
"""

import os
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.properties import PageSetupProperties

warnings.filterwarnings("ignore")

# ----------------------------------------------------------------------------
# CONFIG  -- edit these paths for your machine
# ----------------------------------------------------------------------------
INPUT_DIR       = os.environ.get("INPUT_DIR",  "/Users/phachon/Documents/DKSH/auto-stock-report/input")
OUTPUT_DIR      = os.environ.get("OUTPUT_DIR", "/Users/phachon/Documents/DKSH/auto-stock-report/output")
OUTPUT_NAME     = "Result_Report.xlsx"

FILES = {
    "mb40": "MB52_TH40.XLSX",
    "mb44": "MB52_TH44.XLSX",
    "r40":  "R138_TH40.XLSX",
    "r44":  "R138_TH44.XLSX",
    "pg":   "Product Group.xlsx",          # <-- renamed (space instead of underscore)
}

BUCKET_BINS   = [-np.inf, 30, 90, 180, 365, np.inf]
BUCKET_LABELS = ["0-30", "31-90", "91-180", "181-365", ">365"]
# Display order for the dashboard tables/charts. On the dashboard, R138's
# 'SERVICES' is merged into 'SERVICE' (per request). The raw DATA sheet keeps
# them separate, exactly as the manual report does.
PG_ORDER      = ["ACCESSORIES", "CONSUMABLES", "EQUIPMENT", "SERVICE", "SPARE PARTS"]

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def clean_key(s: pd.Series) -> pd.Series:
    s = s.fillna("").astype(str).str.strip().str.upper()
    s = s.str.replace(r"\.0$", "", regex=True)
    return s.replace("NAN", "")


def upper_strip(s: pd.Series) -> pd.Series:
    """Upper-case + strip a text column. Note: 'SERVICES' is NOT merged into
    'SERVICE' — the manual report keeps them as distinct categories."""
    return s.astype(str).str.strip().str.upper()


def load_inputs():
    p = lambda name: os.path.join(INPUT_DIR, name)
    mb40 = pd.read_excel(p(FILES["mb40"]))
    mb44 = pd.read_excel(p(FILES["mb44"]))
    r40  = pd.read_excel(p(FILES["r40"]))
    r44  = pd.read_excel(p(FILES["r44"]))
    pg   = pd.read_excel(p(FILES["pg"]), sheet_name="Product Group")
    return mb40, mb44, r40, r44, pg


# ----------------------------------------------------------------------------
# GR / Shipper / Profit-center attachment  (tiered keys, per plant)
# ----------------------------------------------------------------------------
def attach_r138(mb, r):
    mb = mb.copy()
    r  = r.copy()
    r["Last GR"] = pd.to_datetime(r["Last GR"], errors="coerce")

    mb["k_full"] = clean_key(mb["Material"]) + "|" + clean_key(mb["Unrestricted"]) + "|" + clean_key(mb["Batch"])
    mb["k_mb"]   = clean_key(mb["Material"]) + "|" + clean_key(mb["Batch"])
    mb["k_mat"]  = clean_key(mb["Material"])
    r["k_full"]  = clean_key(r["Material No."]) + "|" + clean_key(r["Quantity"]) + "|" + clean_key(r["Batch no."])
    r["k_mb"]    = clean_key(r["Material No."]) + "|" + clean_key(r["Batch no."])
    r["k_mat"]   = clean_key(r["Material No."])

    for src_col, new_col in [("Material Group Desc", "Shipper"),
                             ("Profit center", "Profit center")]:
        full = r.drop_duplicates("k_full").set_index("k_full")[src_col]
        mbk  = r.drop_duplicates("k_mb").set_index("k_mb")[src_col]
        val  = mb["k_full"].map(full)
        val  = val.fillna(mb["k_mb"].map(mbk))
        mb[new_col] = val

    # GR Date: full key -> Material+Batch -> Material (most common non-null)
    full = r.drop_duplicates("k_full").set_index("k_full")["Last GR"]
    mbk  = r.drop_duplicates("k_mb").set_index("k_mb")["Last GR"]
    gr   = mb["k_full"].map(full)
    gr   = gr.fillna(mb["k_mb"].map(mbk))
    mat_gr = (r.dropna(subset=["Last GR"])
                .groupby("k_mat")["Last GR"]
                .agg(lambda x: x.value_counts().index[0]))
    gr = gr.fillna(mb["k_mat"].map(mat_gr))
    mb["GR Date"] = gr
    return mb


# ----------------------------------------------------------------------------
# Product Group attachment
# Matches the manual report exactly: Product Group file (by Material) first,
# then RAW R138 'Level 4 Product Group' (by Material) as the only fallback.
# No normalisation (SERVICES stays separate), no supplement, no Material-Group step.
# ----------------------------------------------------------------------------
def attach_product_group(data, pg, r_all):
    data = data.copy()
    data["mk"] = clean_key(data["Material"])

    pg_map = (pg.assign(mk=clean_key(pg["Material"]))
                .drop_duplicates("mk").set_index("mk")["Product Group"].pipe(upper_strip))
    pgv = data["mk"].map(pg_map)

    r_all = r_all.copy()
    r_all["mk"] = clean_key(r_all["Material No."])
    r_all["l4"] = upper_strip(r_all["Level 4 Product Group"])
    l4_map = (r_all.dropna(subset=["Level 4 Product Group"])
                   .drop_duplicates("mk").set_index("mk")["l4"])
    pgv = pgv.fillna(data["mk"].map(l4_map))

    data["Product Group"] = pgv.fillna("Unassigned Group")
    return data


def build_data():
    mb40, mb44, r40, r44, pg = load_inputs()

    for df in (mb40, mb44):
        df["Unrestricted"]       = pd.to_numeric(df["Unrestricted"], errors="coerce")
        df["Value Unrestricted"] = pd.to_numeric(df["Value Unrestricted"], errors="coerce")

    mb40 = mb40[mb40["Unrestricted"] != 0].copy()
    mb44 = mb44[mb44["Unrestricted"] != 0].copy()

    m40 = attach_r138(mb40, r40)
    m44 = attach_r138(mb44, r44)
    data = pd.concat([m40, m44], ignore_index=True)

    r_all = pd.concat([r40, r44], ignore_index=True)
    data = attach_product_group(data, pg, r_all)
    data["Shipper"] = data["Shipper"].fillna("Unassigned Shipper")

    today = pd.Timestamp.today().normalize()
    data["Ageing"] = (today - data["GR Date"]).dt.days
    data["Bucket"] = pd.cut(data["Ageing"], bins=BUCKET_BINS, labels=BUCKET_LABELS)

    out_cols = ["Plant", "Storage location", "Material", "Unrestricted",
                "Value Unrestricted", "Material type", "Material Group",
                "Product Group", "Shipper", "Profit center", "GR Date",
                "Ageing", "Bucket", "Batch"]
    data_out = data[[c for c in out_cols if c in data.columns]].copy()
    return data, data_out


# ----------------------------------------------------------------------------
# Styling primitives
# ----------------------------------------------------------------------------
FONT      = "Segoe UI"
CLR_DARK  = "1F3864"   # navy header
CLR_MED   = "2E5496"   # section title
CLR_SUB   = "D9E1F2"   # subtotal row
CLR_GRAND = "FCE4D6"   # grand-total row
CLR_HL    = "FFF2CC"   # highlight for top-ranked clients (light gold)
WHITE     = "FFFFFF"
CLR_RED   = "C00000"
CLR_GREEN = "548235"
CLR_TEAL  = "1F9E89"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM2 = "#,##0.00;(#,##0.00);-"     # two decimals everywhere
PCT  = "0.00%"

def style_cell(c, *, bold=False, color="000000", size=10, fill=None,
               align="left", numfmt=None, border=True, wrap=False, italic=False):
    c.font = Font(name=FONT, bold=bold, color=color, size=size, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if numfmt:
        c.number_format = numfmt
    if border:
        c.border = BORDER


def merge_band(ws, rng, text, *, fill, color=WHITE, bold=True, size=11,
               align="left", border=False, size_row=None):
    """Merge a range, fill every cell, put text in the top-left."""
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    first_col = "".join(ch for ch in top_left if ch.isalpha())
    first_row = int("".join(ch for ch in top_left if ch.isdigit()))
    last = rng.split(":")[1]
    last_col = "".join(ch for ch in last if ch.isalpha())
    last_row = int("".join(ch for ch in last if ch.isdigit()))
    from openpyxl.utils import column_index_from_string
    for r in range(first_row, last_row + 1):
        for cidx in range(column_index_from_string(first_col),
                           column_index_from_string(last_col) + 1):
            cell = ws.cell(row=r, column=cidx)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name=FONT, bold=bold, color=color, size=size)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if border:
                cell.border = BORDER
    ws[top_left] = text


# ----------------------------------------------------------------------------
# PV DATA dashboard (tables)
# ----------------------------------------------------------------------------
def build_tables_dashboard(ws, data):
    grand_q = data["Unrestricted"].sum()
    grand_v = data["Value Unrestricted"].sum()
    pct = lambda v: (v / grand_v) if grand_v else 0.0

    # Executive summary band
    dead = data[data["Bucket"] == ">365"]
    dead_v = dead["Value Unrestricted"].sum()
    top_client = (dead.groupby("Shipper")["Value Unrestricted"].sum()
                      .sort_values(ascending=False).index[0]) if len(dead) else "-"
    summary = (f"Executive Summary:  Total Inventory Value is {grand_v:,.2f} THB.  "
               f"Dead Stock (>365 Days) accounts for {dead_v:,.2f} THB "
               f"({dead_v/grand_v*100:,.2f}% of total inventory).  "
               f"The primary client driver for dead stock is '{top_client}'.")
    merge_band(ws, "B2:Q4", summary, fill=CLR_DARK, size=11, align="left")
    for r in (2, 3, 4):
        ws.row_dimensions[r].height = 18

    def col_headers(row, first_label):
        for col, txt in zip("BCDE", [first_label, "Quantity", "Stock Value THB.", "%"]):
            style_cell(ws[f"{col}{row}"], bold=True, color=WHITE, size=10,
                       fill=CLR_DARK, align="center")
            ws[f"{col}{row}"] = txt

    def data_row(row, label, q, v, *, indent=False, fill=None, bold=False):
        style_cell(ws[f"B{row}"], bold=bold, fill=fill, align="left")
        ws[f"B{row}"] = ("   " + label) if indent else label
        style_cell(ws[f"C{row}"], bold=bold, fill=fill, align="right", numfmt=NUM2)
        ws[f"C{row}"] = q
        style_cell(ws[f"D{row}"], bold=bold, fill=fill, align="right", numfmt=NUM2)
        ws[f"D{row}"] = v
        style_cell(ws[f"E{row}"], bold=bold, fill=fill, align="right", numfmt=PCT)
        ws[f"E{row}"] = pct(v)

    # 1) Plant Report
    merge_band(ws, "B5:E5", "Plant Report", fill=CLR_MED)
    col_headers(6, "Plant")
    plant_tot = (data.groupby("Plant")[["Unrestricted", "Value Unrestricted"]]
                     .sum().reindex(["TH40", "TH44"]).fillna(0))
    r = 7
    for p in ["TH40", "TH44"]:
        data_row(r, p, plant_tot.loc[p, "Unrestricted"],
                 plant_tot.loc[p, "Value Unrestricted"], fill=CLR_SUB, bold=True)
        r += 1
    data_row(9, "Grand Total", grand_q, grand_v, fill=CLR_GRAND, bold=True)

    # 2) Product Group Report
    merge_band(ws, "B11:E11", "Product Group Report", fill=CLR_MED)
    col_headers(12, "Plant / Product Group")
    r = 13
    for p in ["TH40", "TH44"]:
        sub = data[data["Plant"] == p].copy()
        # On the PV dashboard, merge 'SERVICES' into 'SERVICE' (display only).
        sub["PGdisp"] = sub["Product Group"].replace("SERVICES", "SERVICE")
        data_row(r, p, sub["Unrestricted"].sum(), sub["Value Unrestricted"].sum(),
                 fill=CLR_SUB, bold=True); r += 1
        g = sub.groupby("PGdisp")[["Unrestricted", "Value Unrestricted"]].sum()
        present = [x for x in PG_ORDER if x in g.index] + [x for x in g.index if x not in PG_ORDER]
        for grp in present:
            data_row(r, grp, g.loc[grp, "Unrestricted"], g.loc[grp, "Value Unrestricted"],
                     indent=True); r += 1
    data_row(r, "Grand Total", grand_q, grand_v, fill=CLR_GRAND, bold=True)
    pg_last = r

    # 3) Ageing Report
    age_title = pg_last + 2
    merge_band(ws, f"B{age_title}:E{age_title}", "Ageing Report", fill=CLR_MED)
    hdr = age_title + 1
    col_headers(hdr, "Plant / Ageing")
    r = hdr + 1
    for p in ["TH40", "TH44"]:
        sub = data[data["Plant"] == p]
        data_row(r, p, sub["Unrestricted"].sum(), sub["Value Unrestricted"].sum(),
                 fill=CLR_SUB, bold=True); r += 1
        b = sub.groupby("Bucket", observed=False)[["Unrestricted", "Value Unrestricted"]].sum()
        for bk in BUCKET_LABELS:
            q = b.loc[bk, "Unrestricted"] if bk in b.index else 0
            v = b.loc[bk, "Value Unrestricted"] if bk in b.index else 0
            data_row(r, bk, q, v, indent=True); r += 1
    data_row(r, "Grand Total", grand_q, grand_v, fill=CLR_GRAND, bold=True)

    # 4) + 5) Client tables  (highlight top-N rows + Grand Total)
    def client_block(start_col, title_text, only_dead, highlight_top):
        cols = [get_column_letter(start_col + i) for i in range(5)]
        rng = f"{cols[0]}5:{cols[4]}5"
        merge_band(ws, rng, title_text, fill=CLR_MED)
        for col, txt in zip(cols, ["No.", "Client", "Quantity", "Stock Value THB.", "%"]):
            style_cell(ws[f"{col}6"], bold=True, color=WHITE, size=10, fill=CLR_DARK, align="center")
            ws[f"{col}6"] = txt
        df = data[data["Bucket"] == ">365"] if only_dead else data
        denom = df["Value Unrestricted"].sum()
        g = (df.groupby("Shipper")[["Unrestricted", "Value Unrestricted"]].sum()
               .sort_values("Value Unrestricted", ascending=False))
        rr = 7
        for i, (client, rowv) in enumerate(g.iterrows(), start=1):
            hl = CLR_HL if i <= highlight_top else None
            style_cell(ws[f"{cols[0]}{rr}"], align="center", fill=hl); ws[f"{cols[0]}{rr}"] = i
            style_cell(ws[f"{cols[1]}{rr}"], align="left", fill=hl);   ws[f"{cols[1]}{rr}"] = client
            style_cell(ws[f"{cols[2]}{rr}"], align="right", numfmt=NUM2, fill=hl); ws[f"{cols[2]}{rr}"] = rowv["Unrestricted"]
            style_cell(ws[f"{cols[3]}{rr}"], align="right", numfmt=NUM2, fill=hl); ws[f"{cols[3]}{rr}"] = rowv["Value Unrestricted"]
            style_cell(ws[f"{cols[4]}{rr}"], align="right", numfmt=PCT, fill=hl)
            ws[f"{cols[4]}{rr}"] = (rowv["Value Unrestricted"] / denom) if denom else 0
            rr += 1
        # Grand Total row at the bottom
        style_cell(ws[f"{cols[0]}{rr}"], fill=CLR_GRAND, bold=True, align="center"); ws[f"{cols[0]}{rr}"] = ""
        style_cell(ws[f"{cols[1]}{rr}"], fill=CLR_GRAND, bold=True, align="left");   ws[f"{cols[1]}{rr}"] = "Grand Total"
        style_cell(ws[f"{cols[2]}{rr}"], fill=CLR_GRAND, bold=True, align="right", numfmt=NUM2); ws[f"{cols[2]}{rr}"] = g["Unrestricted"].sum()
        style_cell(ws[f"{cols[3]}{rr}"], fill=CLR_GRAND, bold=True, align="right", numfmt=NUM2); ws[f"{cols[3]}{rr}"] = g["Value Unrestricted"].sum()
        style_cell(ws[f"{cols[4]}{rr}"], fill=CLR_GRAND, bold=True, align="right", numfmt=PCT);  ws[f"{cols[4]}{rr}"] = 1.0

    client_block(7,  "Inventory Clients (THB)", False, 10)
    client_block(13, "Inventory Clients (THB) (>365)", True, 5)

    # widths (wide enough for titles & names)
    widths = {"A": 2, "B": 26, "C": 16, "D": 20, "E": 10, "F": 2,
              "G": 6, "H": 30, "I": 16, "J": 20, "K": 10, "L": 2,
              "M": 6, "N": 30, "O": 16, "P": 20, "Q": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ----------------------------------------------------------------------------
# Executive dashboard (KPI cards + charts)
# ----------------------------------------------------------------------------
def _series_color(series, color):
    series.graphicalProperties = GraphicalProperties(solidFill=color)

def _point_colors(series, colors):
    pts = []
    for i, c in enumerate(colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties = GraphicalProperties(solidFill=c)
        pts.append(dp)
    series.data_points = pts

def build_executive_dashboard(wb, data):
    ws = wb.create_sheet("Executive Dashboard", 0)
    ws.sheet_view.showGridLines = False

    grand_v = data["Value Unrestricted"].sum()
    grand_q = data["Unrestricted"].sum()
    n_sku = data["Material"].nunique()
    n_clients = data["Shipper"].nunique()
    dead = data[data["Bucket"] == ">365"]
    dead_v = dead["Value Unrestricted"].sum()
    dead_pct = dead_v / grand_v if grand_v else 0
    fresh_v = data[data["Bucket"] == "0-30"]["Value Unrestricted"].sum()
    as_of = pd.Timestamp.today().strftime("%d %B %Y")

    # ---- aggregates for charts ----
    plant_v = data.groupby("Plant")["Value Unrestricted"].sum().reindex(["TH40", "TH44"]).fillna(0)
    pg_src = data.copy()
    pg_src["PGdisp"] = pg_src["Product Group"].replace("SERVICES", "SERVICE")
    pg_v = (pg_src.groupby("PGdisp")["Value Unrestricted"].sum()
                  .sort_values(ascending=True))    # sort by value, low -> high
    bucket_v = (data.groupby("Bucket", observed=False)["Value Unrestricted"].sum()
                    .reindex(BUCKET_LABELS).fillna(0))
    top_clients = (data.groupby("Shipper")["Value Unrestricted"].sum()
                       .sort_values(ascending=False).head(10))
    dead_clients = (dead.groupby("Shipper")["Value Unrestricted"].sum()
                        .sort_values(ascending=False).head(10))

    # ---- hidden data sheet for chart sources ----
    cd = wb.create_sheet("_ChartData")
    cd.sheet_state = "hidden"
    def put_block(start_col, header, pairs):
        c0 = start_col
        cd.cell(row=1, column=c0, value=header[0])
        cd.cell(row=1, column=c0 + 1, value=header[1])
        for i, (k, v) in enumerate(pairs, start=2):
            cd.cell(row=i, column=c0, value=str(k))
            cd.cell(row=i, column=c0 + 1, value=float(v))
        return c0
    put_block(1,  ("Plant", "Value"),        list(plant_v.items()))                    # A,B (pie -> %)
    M = lambda s: [(k, round(v / 1e6, 1)) for k, v in s.items()]
    put_block(4,  ("Group", "Value"),        M(pg_v))                                   # D,E (millions)
    put_block(7,  ("Bucket", "Value"),       M(bucket_v))                               # G,H (millions)
    put_block(10, ("Client", "Value"),       M(top_clients[::-1]))                      # J,K (millions, reversed)
    put_block(13, ("Client", "Value"),       M(dead_clients[::-1]))                     # M,N (millions, reversed)

    # ---- column layout ----
    for col in [get_column_letter(i) for i in range(1, 18)]:
        ws.column_dimensions[col].width = 10.5
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["Q"].width = 2.5

    # ---- banner ----
    merge_band(ws, "A1:P1", "DKSH  ·  Stock Ageing Executive Dashboard",
               fill=CLR_DARK, size=18, align="left")
    ws.row_dimensions[1].height = 34
    merge_band(ws, "A2:P2", f"As of {as_of}   ·   All values in THB",
               fill=CLR_MED, size=10, align="left")
    ws.row_dimensions[2].height = 18

    # ---- KPI cards (rows 4-7) ----
    cards = [
        ("B", "E", "TOTAL INVENTORY VALUE", f"{grand_v/1e6:,.2f} M", CLR_DARK),
        ("F", "I", "DEAD STOCK  (>365 DAYS)", f"{dead_v/1e6:,.2f} M  ({dead_pct*100:,.2f}%)", CLR_RED),
        ("J", "M", "DISTINCT SKUs", f"{n_sku:,}", CLR_TEAL),
        ("N", "P", "ACTIVE CLIENTS", f"{n_clients:,}", CLR_GREEN),
    ]
    for c1, c2, label, value, color in cards:
        merge_band(ws, f"{c1}4:{c2}4", label, fill=color, size=9, align="left")
        merge_band(ws, f"{c1}5:{c2}7", value, fill="F2F2F2", color=color, size=18, align="left")
    for r in (4, 5, 6, 7):
        ws.row_dimensions[r].height = 20

    # ---- charts ----
    MFMT = '#,##0.0"M"'   # values already stored in millions -> "149.0M"

    def style_chart(ch, title, h=8.0, w=12.0):
        ch.title = title
        ch.height = h
        ch.width = w
        ch.style = 2

    def show_axes(ch):
        # Force both axes (and their tick labels) to render. Without this,
        # category labels on horizontal bar charts are hidden by the renderer.
        for ax in (ch.x_axis, ch.y_axis):
            ax.delete = False
            ax.tickLblPos = "nextTo"

    def value_labels(ch):
        dl = DataLabelList()
        dl.showVal = True; dl.showCatName = False; dl.showSerName = False
        dl.showLegendKey = False; dl.showPercent = False
        dl.numFmt = MFMT; dl.sourceLinked = False
        ch.dataLabels = dl

    # 1) Plant pie  (legend + percent labels; small slice rotated to the bottom)
    pie = PieChart()
    style_chart(pie, "Inventory Value by Plant", h=8.0, w=11.5)
    labels = Reference(cd, min_col=1, min_row=2, max_row=1 + len(plant_v))
    vals   = Reference(cd, min_col=2, min_row=1, max_row=1 + len(plant_v))
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(labels)
    pie.firstSliceAng = 200          # push the tiny TH40 slice away from the title
    dl = DataLabelList()
    dl.showCatName = True; dl.showPercent = True; dl.showVal = False
    dl.showSerName = False; dl.showLegendKey = False
    pie.dataLabels = dl
    pie.legend.position = "r"        # legend on the right so labels stay light
    _point_colors(pie.series[0], ["5B9BD5", "1F3864"])
    ws.add_chart(pie, "B9")

    # 2) Product group column
    bar = BarChart(); bar.type = "col"
    style_chart(bar, "Inventory Value by Product Group (THB M)", h=8.0, w=12.5)
    cats = Reference(cd, min_col=4, min_row=2, max_row=1 + len(pg_v))
    vals = Reference(cd, min_col=5, min_row=1, max_row=1 + len(pg_v))
    bar.add_data(vals, titles_from_data=True); bar.set_categories(cats)
    bar.legend = None
    show_axes(bar)
    bar.y_axis.numFmt = MFMT; bar.y_axis.majorGridlines = None
    value_labels(bar)
    _point_colors(bar.series[0], ["8FAADC", "5B9BD5", "2E75B6", "2E5496", "1F3864"][:len(pg_v)])
    ws.add_chart(bar, "I9")

    # 3) Ageing column (green -> red)
    age = BarChart(); age.type = "col"
    style_chart(age, "Ageing Profile by Value (THB M)", h=8.0, w=12.5)
    cats = Reference(cd, min_col=7, min_row=2, max_row=1 + len(bucket_v))
    vals = Reference(cd, min_col=8, min_row=1, max_row=1 + len(bucket_v))
    age.add_data(vals, titles_from_data=True); age.set_categories(cats)
    age.legend = None
    show_axes(age)
    age.y_axis.numFmt = MFMT; age.y_axis.majorGridlines = None
    value_labels(age)
    _point_colors(age.series[0], ["70AD47", "A9D18E", "FFD966", "F4B183", "C00000"])
    ws.add_chart(age, "B29")

    # 4) Top 10 clients (horizontal) -- taller so every client name shows
    h1 = BarChart(); h1.type = "bar"
    style_chart(h1, "Top 10 Clients by Value (THB M)", h=10.5, w=14.0)
    cats = Reference(cd, min_col=10, min_row=2, max_row=1 + len(top_clients))
    vals = Reference(cd, min_col=11, min_row=1, max_row=1 + len(top_clients))
    h1.add_data(vals, titles_from_data=True); h1.set_categories(cats)
    h1.legend = None
    show_axes(h1)
    h1.x_axis.numFmt = MFMT; h1.x_axis.majorGridlines = None
    value_labels(h1)
    _series_color(h1.series[0], "2E5496")
    ws.add_chart(h1, "B52")

    # 5) Top 10 dead-stock clients (horizontal, red)
    h2 = BarChart(); h2.type = "bar"
    style_chart(h2, "Top 10 Dead-Stock Clients (>365, THB M)", h=10.5, w=14.0)
    cats = Reference(cd, min_col=13, min_row=2, max_row=1 + len(dead_clients))
    vals = Reference(cd, min_col=14, min_row=1, max_row=1 + len(dead_clients))
    h2.add_data(vals, titles_from_data=True); h2.set_categories(cats)
    h2.legend = None
    show_axes(h2)
    h2.x_axis.numFmt = MFMT; h2.x_axis.majorGridlines = None
    value_labels(h2)
    _series_color(h2.series[0], "C00000")
    ws.add_chart(h2, "I29")

    # ---- page setup: one-page-wide landscape ----
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = "A1:Q74"
    ws.sheet_view.zoomScale = 90
    return ws


# ----------------------------------------------------------------------------
# Plain data sheets
# ----------------------------------------------------------------------------
def write_table(ws, df, *, date_cols=()):
    df = df.copy()
    for dc in date_cols:
        if dc in df.columns:
            df[dc] = pd.to_datetime(df[dc], errors="coerce")
    two_dec = {"Unrestricted", "Value Unrestricted", "Stock value", "Value Unrestricted "}
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        style_cell(c, bold=True, color=WHITE, fill=CLR_DARK, align="center")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            cell = ws.cell(row=i, column=j)
            if isinstance(v, pd.Timestamp):
                cell.value = None if pd.isna(v) else v.to_pydatetime()
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"
            else:
                if pd.isna(v):
                    v = None
                elif isinstance(v, np.integer):
                    v = int(v)
                elif isinstance(v, np.floating):
                    v = float(v)
                cell.value = v
                if col in two_dec and isinstance(v, (int, float)):
                    cell.number_format = NUM2
            cell.font = Font(name=FONT, size=9)
    for j, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(12, min(30, len(str(col)) + 4))
    ws.freeze_panes = "A2"


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Building DATA ...")
    data, data_out = build_data()

    n_nogr = int(data["Bucket"].isna().sum())
    n_unassigned = int((data["Product Group"] == "UNASSIGNED GROUP").sum())
    print(f"  rows: {len(data):,} | No-GR rows: {n_nogr} | Unassigned-Group rows: {n_unassigned}")
    print(f"  grand qty: {data['Unrestricted'].sum():,.3f} | grand value: {data['Value Unrestricted'].sum():,.2f}")

    wb = Workbook()
    # remove default sheet; exec dashboard inserted at index 0
    default_ws = wb.active
    build_executive_dashboard(wb, data)
    wb.remove(default_ws)

    build_tables_dashboard(wb.create_sheet("PV DATA (Dashboard)"), data)
    write_table(wb.create_sheet("DATA"), data_out, date_cols=["GR Date"])
    dead = data_out[data_out["Bucket"] == ">365"].copy()
    write_table(wb.create_sheet("Ageing > 365 D"), dead, date_cols=["GR Date"])

    mb40, mb44, r40, r44, pg = load_inputs()
    for name, df in [("MB52_TH40", mb40), ("MB52_TH44", mb44),
                     ("R138_TH40", r40), ("R138_TH44", r44), ("Product Group", pg)]:
        write_table(wb.create_sheet(name), df,
                    date_cols=[c for c in ("Last GR", "Expiry date") if c in df.columns])

    out_path = os.path.join(OUTPUT_DIR, OUTPUT_NAME)
    wb.save(out_path)
    print(f"Saved -> {out_path}")
    return out_path


if __name__ == "__main__":
    main()